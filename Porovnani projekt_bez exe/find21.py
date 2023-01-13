
class Run:

    def __init__(self, KUS, ID):
        self.KUS = KUS
        self.ID = ID

    def main(self):


        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        import pandas as pd
        from kusovnik_fun_dummy import Kusovnik
        from kusovnik_main_dummy import model
        import numpy as np

        # NACTE DAT Z KUSOVNIKU
        K = Kusovnik(self.KUS)
        M = model(self.ID)

        # COLOURS FOR XLSX
        # https://www.rapidtables.com/web/color/html-color-codes.html
        red   = "E91010"
        green = "10E925"
        orange = "FDC84C"
        tyrkys ="14C1C1"
        gray = "D5D4AE"

        K_next_count, M_next_count = K.count("NEXT"), M.count("NEXT")
        sizeK, sizeM = len(K), len(M)


        ### ZISKANI VSECH KOMPONENT
        KK = []
        MM = []
        Error_list = []
        Error_listM = []
        empty_list   = []
        sb = "NEW.xlsx"
        wb = load_workbook(sb)
        ws = wb.active

        #ws.column_dimensions = 15
        ws.delete_rows(1, ws.max_row + 1)
        Col_labels = ["Komponente", "Menge", "Nazev", "Pos- Material", "Stl./Merkm"]
        ws.append(["POROVNANI", "ID KUSOVNIKU"])
        ws.append(["KUSOVNIK TVL:", M[0]])
        ws.append(["KUSOVNIK maly: ", K[0]])
        ws.append(empty_list)
        ws.append(empty_list)
        ws.append(empty_list)

        print("POCET HODNOT v K: " +str(sizeK))
        print("POCET HODNOT v M: " + str(sizeM))
        print("POCET POLI v K: " + str(K_next_count))
        print("POCET POLI v M: " + str(M_next_count))
        passing = True
        Miss_K = []
        Miss_M = []
        Miss_KK = []
        Miss_MM = []
        Komponente11 = []
        Komponente12=[]
        print("K: " + str(K))
        print("M: "+str(M))
        if passing == True:

            for i in range(1, sizeK):
                if K[i] != "NEXT":
                    KK.append(K[i])
                elif K[i] == "NEXT":
                    if i == sizeK - 1:
                        break
                    else:
                        if str(K[i + 1]) not in str(M):
                            print("K: "+str(K[i + 1])+ " NENI V M")
                            Miss_K.append((K[i + 1]))
                        else:
                            print("In M list finding this  from K: " +str(K[i + 1]))
                            Komponente_posK = M.index(K[i + 1])
                            KPK = Komponente_posK
                            Komponente11.append(M[KPK])

            for t in range(1, sizeM):
                if M[t] != "NEXT":
                    MM.append(M[t])
                elif M[t] == "NEXT":
                    if t == sizeM - 1:
                        break
                    else:
                        if str(M[t + 1]) not in str(K):
                            print("M: "+str(M[t + 1])+ " NENI V K")
                            Miss_M.append((M[t + 1]))
                        else:
                            print("In K list finding this  from M: " +str(M[t + 1]))
                            Komponente_posM = K.index(M[t + 1])
                            KPM = Komponente_posM
                            Komponente12.append(K[KPM])

            #### UNIKATNI KOMPONENTY PRO K
            unique_komponente_K = []
            for ss in Komponente11:
                if ss not in unique_komponente_K:
                    unique_komponente_K.append(ss)

            #### UNIKATNI KOMPONENTY PRO M
            unique_komponente_M = []
            for aa in Komponente12:
                if aa not in unique_komponente_M:
                    unique_komponente_M.append(aa)

            unique_komponente = unique_komponente_K
            print(len(Komponente11))
            print(len(Komponente12))
            print("Miss_M " +str(Miss_M))
            print("Miss_K " +str(Miss_K))
            print(unique_komponente_K)
            print(unique_komponente_M)
            print(len(unique_komponente_K))
            print(len(unique_komponente_M))

        #### ZISKANI POZIC UNIKATNICH KOMPONENT
        uniq_komp_pos_K = []
        uniq_komp_pos_M = []
        M_kompare = []
        K_kompare = []
        M_load = []
        K_load = []
        # Potreba projit celou delku matice unikatnich elementu
        for j in range(len(unique_komponente)):  # pozice unikatniho prvku unique_komponente[j] v matici K
            for i in range(sizeK):  # Potreba projit celou delku matice M
                if K[i] == unique_komponente[j]:
                    uniq_komp_pos_K.append(i)
            for h in range(sizeM):  # Potreba projit celou delku matice M
                if M[h] == unique_komponente[j]:
                    uniq_komp_pos_M.append(h)
            print("pozice pro unikatni znak komponente[" + str(j) + "] vzhledem k listu #K# jsou tyto pozice " + str(
                uniq_komp_pos_K))
            print("pozice pro unikatni znak komponente[" + str(j) + "] vzhledem k listu #M# jsou tyto pozice " + str(
                uniq_komp_pos_M))

            ### ulozeni hodnot z mat. K pro porovnani
            for F in range(len(uniq_komp_pos_K)):
                index = uniq_komp_pos_K[F]
                for index2 in range(index, sizeK):
                    if K[index2] == "NEXT":
                        break

                    else:
                        K_kompare.append(K[index2])
                        K_load.append(K[index2])
                        print("Pro index: " + str(index) + " nalezi hodnota K: " + str(K[index2]))

            ### ulozeni hodnot z mat. M pro porovnani
            for D in range(len(uniq_komp_pos_M)):
                index = uniq_komp_pos_M[D]
                for index2 in range(index, sizeM):
                    if M[index2] == "NEXT":
                        break
                    else:
                        M_kompare.append(M[index2])
                        M_load.append(M[index2])
                        print("Pro index: " + str(index) + " nalezi hodnota M: " + str(M[index2]))
            ### Porovnani
            Size_Err = False
            Miss_Err = False

            for kom in range(len(K_kompare)):

                if len(K_kompare) != len(M_kompare) and K_kompare[kom] not in M_kompare:
                    print("ERR a nenachazi se")
                    Error_list.append("POCET+ERROR")

                elif K_kompare[kom] in M_kompare:
                    print("ok")
                    Error_list.append(" OK")

                elif len(K_kompare) != len(M_kompare):
                    print("Err")
                    print(K_kompare[0])
                    Size_Err = True
                    Error_list.append("POCET")

                else:
                    Miss_Err = True
                    print("nenachazi se")
                    print(K_kompare[0])
                    Error_list.append("ERROR")



            if Size_Err == True:
                ws.append(Col_labels)
                ws.append(M_load)
                ws.append(K_load)
                ws.append(Error_list)
                ws.append(empty_list)
                ws.append(empty_list)
            elif Miss_Err == True:
                ws.append(Col_labels)
                ws.append(M_load)
                ws.append(K_load)
                ws.append(Error_list)
                ws.append(empty_list)
                ws.append(empty_list)
            elif Size_Err and Miss_Err == True:
                ws.append(Col_labels)
                ws.append(M_load)
                ws.append(K_load)
                ws.append(Error_list)
                ws.append(empty_list)
                ws.append(empty_list)
            else:
                ws.append(Col_labels)
                ws.append(M_load)
                ws.append(K_load)
                ws.append(Error_list)
                ws.append(empty_list)
                ws.append(empty_list)

            Error_list.clear()
            wb.save(sb)
            uniq_komp_pos_K.clear()
            uniq_komp_pos_M.clear()
            M_kompare.clear()
            K_kompare.clear()
            M_load.clear()
            K_load.clear()

        NEW_EXCEL = pd.read_excel("NEW.xlsx")
        NEW_EXCEL = np.array(NEW_EXCEL)
        sb = "NEW.xlsx"
        wb = load_workbook(sb)
        ws = wb.active


        NOTIF_MissM=["Toto cislo komponenty se nenachazi v malem KUSOVNIKU"]
        NOTIF_MissK = ["Toto cisla komponenty se nenachazi v TVL"]
        M_missing = []
        K_missing = []
        if len(Miss_M) > 0:

            for i in range(len(Miss_M)):
                miss_index = M.index(Miss_M[i])
                for j in range(miss_index, sizeM):
                    if str(M[j]) != "NEXT":
                        M_missing.append(M[j])
                    elif str(M[j]) == "NEXT":
                        ws.append(empty_list)
                        ws.append(empty_list)
                        ws.append(NOTIF_MissM)
                        ws.append(Col_labels)
                        ws.append(M_missing)
                        M_missing.clear()
                        break
            wb.save(sb)

        if len(Miss_K) > 0:
            for c in range(len(Miss_K)):
                miss_i = K.index(Miss_K[c])
                for z in range(miss_i, sizeK):
                    if str(K[z]) != "NEXT":
                        K_missing.append(K[z])
                    elif str(K[z]) == "NEXT":
                        ws.append(empty_list)
                        ws.append(empty_list)
                        ws.append(NOTIF_MissK)
                        ws.append(Col_labels)
                        ws.append(K_missing)
                        K_missing.clear()
                        break
            wb.save(sb)
        new_row, new_col = NEW_EXCEL.shape
        for LL in range(new_col):
            for PP in range(new_row):
                if str(NEW_EXCEL[PP][LL]) == " OK":
                    ws[PP + 2][LL].fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
                elif str(NEW_EXCEL[PP][LL]) == "ERROR":
                    ws[PP + 2][LL].fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
                elif str(NEW_EXCEL[PP][LL]) == "POCET+ERROR":
                    ws[PP + 2][LL].fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
                elif (NEW_EXCEL[PP][LL]) in (unique_komponente):
                    ws[PP + 2][LL].fill = PatternFill(start_color=tyrkys, end_color=tyrkys, fill_type="solid")
                elif str(NEW_EXCEL[PP][LL]) in Col_labels:
                    ws[PP + 2][LL].fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid")
                else:
                    pass
        wb.save(sb)
